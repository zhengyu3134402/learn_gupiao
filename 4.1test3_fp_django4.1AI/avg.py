import time
import openpyxl
from lib.second_conn_sqlite3 import *


result2_list = []

def from_database2_take_data(clear_name):



    session = second_conn_sqlite3.make_session()

    all_data = list(session.query(Second).filter_by(name=clear_name))
    # result2_list.append((all_data, clear_name))

    session.close()
    return  all_data


result3_list = []
def write_to_excel(all_data):
    # print(all_data)
    avg_list = []

    sp2_sp0_list = []
    kp2_sp1_list = []
    zg2_sp1_list = []
    zd2_sp1_list = []
    sp2_sp1_list = []
    kp3_sp1_list = []
    zg3_sp1_list = []
    zd3_sp1_list = []
    sp3_sp1_list = []
    kp4_sp1_list = []
    zg4_sp1_list = []
    zd4_sp1_list = []
    sp4_sp1_list = []
    kp5_sp1_list = []
    zg5_sp1_list = []
    zd5_sp1_list = []
    sp5_sp1_list = []
    kp6_sp1_list = []
    zg6_sp1_list = []
    zd6_sp1_list = []
    sp6_sp1_list = []

    for obj in all_data:
        sp2_sp0_list.append(obj.sp2_sp0)
        kp2_sp1_list.append(obj.kp2_sp1)
        zg2_sp1_list.append(obj.zg2_sp1)
        zd2_sp1_list.append(obj.zd2_sp1)
        sp2_sp1_list.append(obj.sp2_sp1)
        kp3_sp1_list.append(obj.kp3_sp1)
        zg3_sp1_list.append(obj.zg3_sp1)
        zd3_sp1_list.append(obj.zd3_sp1)
        sp3_sp1_list.append(obj.sp3_sp1)
        kp4_sp1_list.append(obj.kp4_sp1)
        zg4_sp1_list.append(obj.zg4_sp1)
        zd4_sp1_list.append(obj.zd4_sp1)
        sp4_sp1_list.append(obj.sp4_sp1)
        kp5_sp1_list.append(obj.kp5_sp1)
        zg5_sp1_list.append(obj.zg5_sp1)
        zd5_sp1_list.append(obj.zd5_sp1)
        sp5_sp1_list.append(obj.sp5_sp1)
        kp6_sp1_list.append(obj.kp6_sp1)
        zg6_sp1_list.append(obj.zg6_sp1)
        zd6_sp1_list.append(obj.zd6_sp1)
        sp6_sp1_list.append(obj.sp6_sp1)


    result3_list.append([all_data[0].name,
                         sum(sp2_sp0_list) / len(sp2_sp0_list),
                         sum(kp2_sp1_list) / len(kp2_sp1_list),
                         sum(zg2_sp1_list) / len(zg2_sp1_list),
                         sum(zd2_sp1_list) / len(zd2_sp1_list),
                         sum(sp2_sp1_list) / len(sp2_sp1_list),
                         sum(kp3_sp1_list) / len(kp3_sp1_list),
                         sum(zg3_sp1_list) / len(zg3_sp1_list),
                         sum(zd3_sp1_list) / len(zd3_sp1_list),
                         sum(sp3_sp1_list) / len(sp3_sp1_list),
                         sum(kp4_sp1_list) / len(kp4_sp1_list),
                         sum(zg4_sp1_list) / len(zg4_sp1_list),
                         sum(zd4_sp1_list) / len(zd4_sp1_list),
                         sum(sp4_sp1_list) / len(sp4_sp1_list),

                         sum(kp5_sp1_list) / len(kp5_sp1_list),
                         sum(zg5_sp1_list) / len(zg5_sp1_list),
                         sum(zd5_sp1_list) / len(zd5_sp1_list),
                         sum(sp5_sp1_list) / len(sp5_sp1_list),
                         sum(kp6_sp1_list) / len(kp6_sp1_list),
                         sum(zg6_sp1_list) / len(zg6_sp1_list),
                         sum(zd6_sp1_list) / len(zd6_sp1_list),
                         sum(sp6_sp1_list) / len(sp6_sp1_list),
                         len(sp2_sp0_list)])


def make_excel():



    filepath = os.getcwd() + '\\avg.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['技术指标', '次日收盘 / 昨日收盘(平均数)', '次日开盘 / 当日收盘(平均数)', '次日最高 / 当日收盘(平均数)',
               '次日最低 / 当日收盘(平均数)', '次日收盘 / 当日收盘(平均数)', '第2日开盘 / 当日收盘(平均数)',
               '第2日最高 / 当日收盘(平均数)', '第2日最低 / 当日收盘(平均数)', '第2日收盘 / 当日收盘(平均数)',
               '第3日开盘 / 当日收盘(平均数)', '第3日最高 / 当日收盘(平均数)', '第3日最低 / 当日收盘(平均数)',
               '第3日收盘 / 当日收盘(平均数)', '第4日开盘 / 当日收盘(平均数)', '第4日最高 / 当日收盘(平均数)',
               '第4日最低 / 当日收盘(平均数)', '第4日收盘 / 当日收盘(平均数)', '第5日开盘 / 当日收盘(平均数)',
               '第5日最高 / 当日收盘(平均数)', '第5日最低 / 当日收盘(平均数)', '第5日收盘 / 当日收盘(平均数)',
               '个数'])
    wb.save(filepath)











if __name__ == '__main__':


# 多进程===================================================================
    from multiprocessing import Pool

    make_excel()

    t1 = time.time()
    p = Pool(5)



    with open('jishuzhibiao.txt', 'r') as f:
        s = f.readlines()
        for name in s:
            clear_name = name.strip()

            result = p.apply_async(from_database2_take_data, args=(clear_name,), callback=write_to_excel)

        p.close()
        p.join()
    # print(result3_list)
    filepath = os.getcwd() + '\\avg.xlsx'
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    for i in result3_list:
        # print(i)
        ws.append(i)
    wb.save(filepath)

    t2=time.time()
    result_time = t2-t1
    print(';;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;')
    print('计算平均数花了%r秒, 请查看avg.xlsx表格！！'%(result_time))
    print(';;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;')
# 多进程===================================================================

# 单线程 ================================================================
# def from_database2_take_data(clear_name):
#
#
#
#     session = second_conn_sqlite3.make_session()
#
#     all_data = list(session.query(Second).filter_by(name=clear_name))
#     write_to_excel(all_data, clear_name)
#     session.close()
#
#
#
#
# def write_to_excel(all_data, clear_name):
#
#     avg_list = []
#
#     filepath = os.getcwd() + '\\avg.xlsx'
#     wb = openpyxl.load_workbook(filepath)
#     ws = wb.active
#     for obj in all_data:
#
#         avg_list.append(obj.sp2_sp1)
#
#     ws.append([clear_name, sum(avg_list)/len(avg_list), len(avg_list)])
#     wb.save(filepath)
#
# def make_excel():
#
#     filepath = os.getcwd() + '\\avg.xlsx'
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.append(['技术指标', '平均数', '个数'])
#     wb.save(filepath)
#
#
#
#
# def computed_avg():
#     with open('jishuzhibiao.txt', 'r') as f:
#
#         s = f.readlines()
#         for name in s:
#             clear_name = name.strip()
#             from_database2_take_data(clear_name)
#
# t1 = time.time()
# make_excel()
# computed_avg()
# t2 = time.time()
#
#
# result_time = t2-t1
#
# print(';;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;')
# print('总共花了%r秒, 请查看avg.xlsx表格！！'%(result_time))
# print(';;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;')
# 单线程 ================================================================