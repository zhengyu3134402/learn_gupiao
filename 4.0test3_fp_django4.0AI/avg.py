import time
import openpyxl
from lib.second_conn_sqlite3 import *


result2_list = []

def from_database2_take_data(clear_name):



    session = second_conn_sqlite3.make_session()

    all_data = list(session.query(Second).filter_by(name=clear_name))
    # result2_list.append((all_data, clear_name))

    session.close()
    return  all_data



result3_list = []
def write_to_excel(all_data):
    # print(all_data)
    avg_list = []

    # print(1)

    for obj in all_data:


        avg_list.append(obj.sp2_sp1)

    result3_list.append([all_data[0].name, sum(avg_list)/len(avg_list), len(avg_list)])


def make_excel():

    filepath = os.getcwd() + '\\avg.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['技术指标', '平均数', '个数'])
    wb.save(filepath)











if __name__ == '__main__':


# 多进程===================================================================
    from multiprocessing import Pool

    make_excel()

    t1 = time.time()
    p = Pool(5)



    with open('jishuzhibiao.txt', 'r') as f:
        s = f.readlines()
        for name in s:
            clear_name = name.strip()

            result = p.apply_async(from_database2_take_data, args=(clear_name,), callback=write_to_excel)

        p.close()
        p.join()
    # print(result3_list)
    filepath = os.getcwd() + '\\avg.xlsx'
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    for i in result3_list:
        # print(i)
        ws.append(i)
    wb.save(filepath)

    t2=time.time()
    result_time = t2-t1
    print(';;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;')
    print('计算平均数花了%r秒, 请查看avg.xlsx表格！！'%(result_time))
    print(';;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;')
# 多进程===================================================================

# 单线程 ================================================================
# def from_database2_take_data(clear_name):
#
#
#
#     session = second_conn_sqlite3.make_session()
#
#     all_data = list(session.query(Second).filter_by(name=clear_name))
#     write_to_excel(all_data, clear_name)
#     session.close()
#
#
#
#
# def write_to_excel(all_data, clear_name):
#
#     avg_list = []
#
#     filepath = os.getcwd() + '\\avg.xlsx'
#     wb = openpyxl.load_workbook(filepath)
#     ws = wb.active
#     for obj in all_data:
#
#         avg_list.append(obj.sp2_sp1)
#
#     ws.append([clear_name, sum(avg_list)/len(avg_list), len(avg_list)])
#     wb.save(filepath)
#
# def make_excel():
#
#     filepath = os.getcwd() + '\\avg.xlsx'
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.append(['技术指标', '平均数', '个数'])
#     wb.save(filepath)
#
#
#
#
# def computed_avg():
#     with open('jishuzhibiao.txt', 'r') as f:
#
#         s = f.readlines()
#         for name in s:
#             clear_name = name.strip()
#             from_database2_take_data(clear_name)
#
# t1 = time.time()
# make_excel()
# computed_avg()
# t2 = time.time()
#
#
# result_time = t2-t1
#
# print(';;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;')
# print('总共花了%r秒, 请查看avg.xlsx表格！！'%(result_time))
# print(';;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;;')
# 单线程 ================================================================